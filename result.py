
from openpyxl import Workbook, load_workbook 
wb=load_workbook('tests/test1.xlsx')
ws=wb.active
max_row = ws.max_row
print(max_row)
g = 0

for i in range (2,max_row + 1):
   
        hours=ws["B" + str(i)].value
        rate=ws["C"+str(i)].value
         
        if (type(hours)!=str and type(rate)!=str):   
            salary = hours * rate
            (ws["D"+str(i)].value) = salary
            print(salary)

for i in range (2,max_row + 1): 
      salary=ws["D"+str(i)].value
      
      if salary == int and salary >= 3000:
        g + 1
        (ws["E"+str(i)].value) = g
        print(g)
          
wb.save('finalResult.xlsx')
wb.close() 
                
